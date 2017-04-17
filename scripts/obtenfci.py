# -*- coding: utf-8 -*-
import signal

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
import time
import logging
import openpyxl
import zipfile
from datetime import datetime
from os.path import basename
import win32com.client as win32
import pythoncom
import shutil, os
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By


# Funciones Utils
from selenium.webdriver.support.wait import WebDriverWait


def es_aval(row, hp):
    global aval
    if hp.cell(row=row, column=6).value == 'Aval PM':
        aval = True
    elif hp.cell(row=row, column=6).value == 'Amont PM':
        aval = False
    return aval


def calculo_tipo(row, hp):
    global tipo
    if hp.cell(row=row, column=20).value == 'SIMPLE':
        tipo = 'SPL'
    elif hp.cell(row=row, column=20).value == 'COMPLEXE':
        tipo = 'CPL'
    elif hp.cell(row=row, column=20).value == 'STRUCTURANTE':
        tipo = 'STR'
    return tipo


def calculo_fechas(tipo, col, hp):
    global date
    if tipo == 'SPL':
        date = hp.cell(row=1, column=col).value
    elif tipo == 'CPL':
        date = hp.cell(row=2, column=col).value
    elif tipo == 'STR':
        date = hp.cell(row=3, column=col).value
    date = date.strftime('%d/%m/%Y')
    return date


def calculo_arquetas(row, hp):
    global arquetas
    if hp.cell(row=row, column=25).value != 0 and hp.cell(row=row, column=26).value == 0:
        arquetas = 'gc'
    elif hp.cell(row=row, column=25).value != 0 and hp.cell(row=row, column=26).value != 0:
        arquetas = 'gc+p'
    elif hp.cell(row=row, column=25).value == 0 and hp.cell(row=row, column=26).value != 0:
        arquetas = 'p'
    return arquetas


def calculo_numel(row, hp):
    global num_el
    if hp.cell(row=row, column=727).value == None:
        num_el = '0'
    else:
        a = hp.cell(row=row, column=727).value
        num_el = str(a)
    return num_el


def calculo_calles(row, hp):
    global calles
    try:
        calles = []
        if hp.cell(row=row, column=7).value == 'D2 IMB':
            calles.append(hp.cell(row=row, column=12).value + ' ' + hp.cell(row=row, column=13).value)
        if hp.cell(row=row, column=14).value != None:
            otras_calles = hp.cell(row=row, column=14).value
            otras_calles = otras_calles.split(sep='/')
            for c in otras_calles:
                calles.append(c)
    except:
        raise Exception('No tiene calles', hp.cell(row=row, column=1).value)
    return calles


def calculo_formulario(ciudad, tp, hf, client):
    global formulario
    for row in hf.iter_rows(min_row=1, max_col=1, max_row=hf.max_row):
        for celda in row:
            if celda.value == ciudad:
                column = 9
                if tp == 'CPL':
                    column = 10
                elif tp == 'STR':
                    column = 11
                if client == 'SC6': column += 1
                formulario = hf.cell(row=celda.row, column=column).value
    return formulario


def cargar_datos_excel(client):
    # Carga de todos los datos necesarios del excel y los mete en un diccionario de dosieres
    global doc
    print('Carga de todos los datos necesarios del excel y los mete en un diccionario de dosieres')
    # doc = openpyxl.load_workbook('SuiviJRU.xlsx')

    try:
        if os.name == 'nt':
            # doc = openpyxl.load_workbook('SuiviJRU.xlsm', data_only=True)
            if client == 'SC1':
                doc = openpyxl.load_workbook(r'Z:/03-PRODUCCION/0.CAFT/SC1/PRODUCCIÓN/Tab Suivi Prod/SC1 TSP 2017.xlsm',
                                             data_only=True)
            elif client == 'SC00':
                doc = openpyxl.load_workbook(
                    r'Z:/03-PRODUCCION/0.CAFT/SC00/PRODUCCIÓN/Tab Suivi Prod/SC00 TSP 2017.xlsm',
                    data_only=True)
            elif client == 'SC6':
                doc = openpyxl.load_workbook(r'Z:/03-PRODUCCION/0.CAFT/SC6/PRODUCCIÓN/Tab Suivi Prod/SC6 TSP 2017.xlsm',
                                             data_only=True)
            elif client == 'SC4':
                doc = openpyxl.load_workbook(r'Z:/03-PRODUCCION/0.CAFT/SC4/PRODUCCIÓN/Tab Suivi Prod/SC4 TSP 2017.xlsm',
                                             data_only=True)
        else:
            doc = openpyxl.load_workbook(r'/home/ubuntu/3id2plus/SuiviJRU.xlsx', data_only=True)
            # doc = openpyxl.load_workbook(r'/home/ubuntu/nas/NAS/03-PRODUCCION/0.CAFT/SC1/PRODUCCIÓN/Tab Suivi Prod/suivi prod general SC1 practica-JRU.xlsx', data_only=True)
        doc.get_sheet_names()
        hoja_principal = doc.get_sheet_by_name('Tab Suivi Prod')
        hoja_formulario = doc.get_sheet_by_name('DEX & PIT')
    except:
        raise Exception('El archivo no existe o no tiene esas pestañas', 'Error: ')
    dosieres = dict()

    try:
        for row in hoja_principal.iter_rows(min_row=1, max_col=1, max_row=hoja_principal.max_row):
            for celda in row:
                fci_cell = hoja_principal.cell(row=celda.row, column=40).value
                sacar_fci_simple = hoja_principal.cell(row=celda.row, column=20).value == 'SIMPLE' and \
                                   (fci_cell == None or fci_cell[1] == '$') and celda.row != 1565
                sacar_fci_cplstr = (hoja_principal.cell(row=celda.row, column=20).value == 'COMPLEXE' or
                                    hoja_principal.cell(row=celda.row, column=20).value == 'STRUCTURANTE') \
                                   and (fci_cell == None or fci_cell[1] == '$') and \
                                   hoja_principal.cell(row=celda.row, column=50).value == 'CTRL OK'

                if (sacar_fci_cplstr or sacar_fci_simple) and \
                        (hoja_principal.cell(row=celda.row, column=15).value != 'en attente' and
                                 hoja_principal.cell(row=celda.row, column=15).value != 'Not FCI') and \
                                hoja_principal.cell(row=celda.row, column=3).value == '3ID2+':
                    #                   formulario en pestaña dex & pit y el numero fi aax   y 727 para la capacidad el
                    dosier = {
                        'nombre': celda.value,
                        'ciudad': hoja_principal.cell(row=celda.row, column=16).value,
                        'otras_ciudades': hoja_principal.cell(row=celda.row, column=17).value,
                        'es_aval': es_aval(celda.row, hoja_principal),
                        'tipo': calculo_tipo(celda.row, hoja_principal),
                        'es_1ca': hoja_principal.cell(row=celda.row, column=19).value == '1er CA',
                        'IPE_PM': hoja_principal.cell(row=celda.row, column=8).value,
                        'ref_1era_PM': hoja_principal.cell(row=celda.row, column=15).value,
                        'num_EL': calculo_numel(celda.row, hoja_principal),
                        'cliente': client,
                        'solo_arquetas': calculo_arquetas(celda.row, hoja_principal),  # gc, gc+p, p
                        'calles': calculo_calles(celda.row, hoja_principal),
                        'ref_cli': hoja_principal.cell(row=celda.row, column=79).value,
                        'row': celda.row
                    }
                    if fci_cell is not None:
                        if fci_cell[1] == '$':
                            dosier['fci_anterior'] = fci_cell
                    dosier['formulario'] = calculo_formulario(dosier['ciudad'], dosier['tipo'], hoja_formulario, client)
                    dosier['date_ini'] = calculo_fechas(dosier['tipo'], 4, hoja_principal)
                    dosier['date_fin'] = calculo_fechas(dosier['tipo'], 5, hoja_principal)
                    dosieres[dosier['nombre']] = dosier
    except:
        pass
    finally:
        doc = None
    return dosieres


############################################ EJEMPLO DOSIER ###########################
# dosier : {
#     'nombre': 'Les1303'
#     'cliente': 'SC1',
#     'tipo': 'SPL',
#     'ciudad': 'COLOMBES',
#     'es_aval': True,
#     'es_1ca': True,
#     'solo_arquetas': True
#     'IPE_PM': 'FI-63463-12312',
#     'ref_1era_PM': 'F1241231233',
#     'num_el': '34',
#     'date_ini': '05/12/2016',
#     'date_fin': '02/01/2016',
#     'calles': ['rue del percebe', 'calle street'],
#     'ref_cli': 'SC1_IMB_93045_C_00X1',
#     'fci': 'F234124123'
#     'row': 345
# }
######################################################################################

def set_up_browser():
    # set up browser
    print('set up browser')
    # Windows
    if os.name == 'nt':
        ############################# FIREFOX ###########################################
        binary = FirefoxBinary(r'C:\Program Files (x86)\Mozilla Firefox\firefox.exe')
        # Crear un profile para omitir los cuadros de diálogo al descargar archivos
        fp = webdriver.FirefoxProfile()
        fp.set_preference('browser.download.folderList', 2)
        fp.set_preference('browser.download.manager.showWhenStarting', False)
        fp.set_preference('browser.download.dir', os.getcwd())
        fp.set_preference('browser.helperApps.neverAsk.openFile', 'application/zip')
        fp.set_preference('browser.helperApps.neverAsk.saveToDisk','application/zip')
        fp.set_preference('browser.helperApps.alwaysAsk.force', False)
        browser = webdriver.Firefox(firefox_binary=binary, firefox_profile=fp)

        ############################ PHANTOMJS ##########################################
        # path = r'C:\Users\josko\PycharmProjects\josko\scripts\phantomjs-2.1.1-windows\bin\phantomjs.exe'
        # browser = webdriver.PhantomJS(executable_path=path)
    # Linux
    else:
        browser = webdriver.PhantomJS()

    browser.get('https://dro.orange-business.com/authentification?target=https://espaceclient.orange-business.com/group'
                '/divop/home?codeContexte=ece_divop&TYPE=33554433&REALMOID=06-00006a03-1ec3-1184-b5ad-5e0e0a63d064&'
                'GUID=&SMAUTHREASON=0&METHOD=GET&SMAGENTNAME=-SM-hCJfMsHRC8Nvudq1lQDbznywak%2fg%2bYsE6nklHqOsEk8XmYpdaqDy'
                'ezDHzkpWx6GU&TARGET=-SM-https%3a%2f%2fespaceclient%2eorange--business%2ecom%2fgroup%2fdivop%2fhome')
    return browser


def login(browser, client):
    print('login')
    elem = browser.find_element_by_id("username")
    elem2 = browser.find_element_by_id("password")
    if client == 'SC1':
        elem.send_keys("michael.yniesta")
        time.sleep(1)
        elem2.send_keys("Scopelec92!" + Keys.RETURN)
    elif client == 'SC00':
        elem.send_keys("LLWQ3863")
        time.sleep(1)
        elem2.send_keys("Reunion974*" + Keys.RETURN)
    elif client == 'SC6':
        elem.send_keys('lcosma@groupe-scopelec.fr')
        time.sleep(1)
        elem2.send_keys("Scopelec31*" + Keys.RETURN)
    elif client == 'SC4':
        elem.send_keys('faiza.kheda')
        time.sleep(1)
        elem2.send_keys("Scopelec03/" + Keys.RETURN)
    time.sleep(4)


def boutique_operations(browser, d):
    'Rellena formulario con los datos del dosier para obtener el número FCI'
    print('boutique operations ' + d['nombre'])
    dos_type = d['tipo']
    browser.get('https://espaceclient.orange-business.com/group/divop/boutique-operateurs')
    time.sleep(3)
    main_window = browser.current_window_handle

    browser.switch_to_frame(browser.find_element_by_id('ece_iframe'))
    a = browser.find_element_by_css_selector('.sfci_box > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > '
                                             'table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2)'
                                             ' > select:nth-child(1) > option:nth-child(7)')
    a.click()
    time.sleep(1)
    # Elegir operation, esperamos para que se carguen las opciones
    if dos_type == 'SPL':
        b = browser.find_element_by_css_selector('.sfci_box > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > '
                                                 'table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(2) '
                                                 '> select:nth-child(1) > option:nth-child(8)')
    elif dos_type == 'CPL':
        b = browser.find_element_by_css_selector('.sfci_box > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > '
                                                 'table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(2)'
                                                 ' > select:nth-child(1) > option:nth-child(7)')
    elif dos_type == 'STR':
        b = browser.find_element_by_css_selector('.sfci_box > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1)'
                                                 ' > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > '
                                                 'td:nth-child(2) > select:nth-child(1) > option:nth-child(4)')

    b.click()
    c = browser.find_element_by_css_selector('.sfci_box > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) '
                                             '> table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(3) > td:nth-child(2)'
                                             ' > select:nth-child(1) > option:nth-child(2)')
    c.click()
    time.sleep(1)
    browser.find_element_by_css_selector('a.sfci_blackb:nth-child(5)').click()  # pulsamos en 'deposer a partir d'une...
    # se abre una nueva ventana y hay que elegir el formulario que corresponda
    time.sleep(1)
    # signin_window_handle = None
    # while not signin_window_handle:
    #     for handle in browser.window_handles:
    #         if handle != main_window_handle:
    #             signin_window_handle = handle
    #             break
    WebDriverWait(browser, 15).until(EC.number_of_windows_to_be(2))
    signin_window_handle = [window for window in browser.window_handles if window != main_window][0]
    browser.switch_to.window(signin_window_handle)
    formulario = d['formulario']
    time.sleep(5)
    row_text = WebDriverWait(browser, 20).until(
        EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '" + formulario + "')]")))
    time.sleep(3)
    row_parent = row_text.find_element_by_xpath('..')
    clickable_button = row_parent.find_element_by_xpath('.//input')
    clickable_button.click()
    valider_button = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.sfci_blackb:nth-child(2)')))
    valider_button.click()
    WebDriverWait(browser, 15).until(EC.number_of_windows_to_be(1))
    browser.switch_to.window(main_window)
    time.sleep(1)

    # Volvemos a la pagina anterior y aparece un formulario en el que hay que rellenar algunos campos
    es_aval = d['es_aval']
    aval_primera = d['es_1ca']
    ref_1era_PM = d['ref_1era_PM']
    IPE_PM = d['IPE_PM']
    solo_arquetas = d['solo_arquetas']
    fecha_ini = d['date_ini']
    fecha_fin = d['date_fin']
    ref_cli = d['ref_cli']
    calles = d['calles']
    try:
        d['calles'][0]
    except:
        raise Exception('No tiene calles')
    # Si esta en phantomjs tiene que entrar en el iframe
    if browser.name == 'phantomjs':
        browser.switch_to_frame('ece_iframe')
    time.sleep(1)
    commande = browser.find_element_by_css_selector(
        '#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_pm_td_input '
        '> select:nth-child(1)')
    commande.click()
    time.sleep(1)

    # Para dosier Simple
    if dos_type == 'SPL':
        aval1 = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:ref_cde_aval_pm_td'
                                                     '_input > input:nth-child(1)')
        aval1.clear()
        time.sleep(1)
        aval2 = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:ref_ipe_td_input '
                                                     '> input:nth-child(1)')
        aval2.clear()
        time.sleep(1)
        if es_aval:
            # opción marcada por defecto. No modificar el select, solo los 2 campos siguientes
            browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_pm_td_input'
                                                 ' > select:nth-child(1) > option:nth-child(1)').click()
            aval_choice = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde'
                                                               '_pm_td_input > select:nth-child(1) > option:nth-child(3)')
            aval_choice.click()
            aval1.send_keys(ref_1era_PM)
            aval2.send_keys(IPE_PM)

        elif not es_aval:
            time.sleep(1)
            browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_'
                                                 'pm_td_input > select:nth-child(1) > option:nth-child(1)').click()
            amont_choice = browser.find_element_by_css_selector(
                '#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_'
                'pm_td_input > select:nth-child(1) > option:nth-child(2)')
            amont_choice.click()  # \/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_pm_td_input > select:nth-child(1) > option:nth-child(2)

        time.sleep(2)
        browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                             'concerne > option:nth-child(1)').click()
        # Si solo arquetas se deja por defecto GC, si arquetas + poteaux se pone GC et apus aeris
        select_postes = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                                             'concerne > option:nth-child(2)')
        if solo_arquetas == 'gc+p':
            select_postes = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                                                 'concerne > option:nth-child(4)')
        elif solo_arquetas == 'p':
            select_postes = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                                                 'concerne > option:nth-child(3)')
        select_postes.click()

        time.sleep(1)
        # Fechas de simple
        form_fecha_ini = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contenu_decl_trvx\/gcblo\:'
                                                              'date_debut_trvx')
        form_fecha_fin = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contenu_decl_trvx\/gcblo\:'
                                                              'date_fin_trvx_td_input > input:nth-child(1)')
        time.sleep(1)
        form_fecha_ini.clear()
        form_fecha_ini.send_keys(fecha_ini)
        time.sleep(1)
        form_fecha_fin.clear()
        form_fecha_fin.send_keys(fecha_fin)

        time.sleep(1)
        #     Annadimos las calles que pasan por el recorrido
        for i in range(len(calles)):
            calle_form = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contenu_decl_trvx\/gcblo\:'
                                                              'arteres_princ\[' + str(
                i + 1) + '\]_td_input > input:nth-child(1)')
            time.sleep(1)
            calle_form.clear()
            calle_form.send_keys(calles[i])
            if i < len(calles) - 1:
                time.sleep(1)
                add_button = browser.find_element_by_css_selector(
                    '#\/com\:commande\/gcblo\:contenu_decl_trvx\/gcblo\:arteres_princ\[1\]_td_label1 > a:nth-child(1) > '
                    'img:nth-child(1)')
                add_button.click()

    # Si el dosier es complejo o estructurante
    # af = zone de commande
    # bf = IPE de PM
    # cf = type de commande
    # df = Num el
    # ef = cable mixte
    # ff = FCI del primer
    if dos_type == 'CPL' or dos_type == 'STR':
        bf = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:ref_ipe_td_input > '
                                                  'input:nth-child(1)')
        bf.clear()
        time.sleep(1)
        df = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:taille_pm_td_input'
                                                  ' > input:nth-child(1)')
        df.clear()
        time.sleep(1)
        browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:cab_mixte_td_input > '
                                             'select:nth-child(1) > option:nth-child(1)').click()
        ff = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:ref_cde_aval_pm_td_'
                                                  'input > input:nth-child(1)')
        ff.clear()
        time.sleep(1)
        browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:type_cde_td_input'
                                             ' > select:nth-child(1) > option:nth-child(1)').click()
        browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_pm_'
                                             'td_input > select:nth-child(1) > option:nth-child(1)').click()
        time.sleep(1)

        if es_aval and not aval_primera:
            af = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_pm_'
                                                      'td_input > select:nth-child(1) > option:nth-child(2)')
            af.click()
            time.sleep(1)
            bf.send_keys(IPE_PM)
            time.sleep(1)
            cf = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:type_cde_td_input'
                                                      ' > select:nth-child(1) > option:nth-child(3)')
            cf.click()
            time.sleep(1)

            ff.send_keys(ref_1era_PM)

        if es_aval and aval_primera:
            num_el = d['num_EL']
            af = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:'
                                                      'zone_cde_pm_td_input > select:nth-child(1) > option:nth-child(2)')
            af.click()
            time.sleep(1)
            bf.send_keys(IPE_PM)
            cf = browser.find_element_by_css_selector(
                '#\/com\:commande\/gcblo\:contexte_com\/gcblo\:type_cde_td_input > '
                'select:nth-child(1)')
            cf.click()
            time.sleep(1)
            c1 = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:type_cde_td_input'
                                                      ' > select:nth-child(1) > option:nth-child(2)')
            c1.click()
            df.send_keys(num_el)
            time.sleep(1)
            ef = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:cab_mixte_td_input'
                                                      ' > select:nth-child(1)')
            ef.click()
            time.sleep(1)
            browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:cab_mixte_td_input >'
                                                 ' select:nth-child(1) > option:nth-child(3)').click()

            time.sleep(1)

        if not es_aval:
            time.sleep(1)
            af = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:zone_cde_pm_td_'
                                                      'input > select:nth-child(1) > option:nth-child(3)')
            af.click()
            time.sleep(1)
            cf = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:type_cde_td_input'
                                                      ' > select:nth-child(1)')
            cf.click()
            time.sleep(1)
            c1 = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:contexte_com\/gcblo\:type_cde_td_input'
                                                      ' > select:nth-child(1) > option:nth-child(1)')
            c1.click()

        time.sleep(2)
        # Si solo arquetas se deja por defecto GC, si arquetas + poteaux se pone GC et apus aeris
        browser.find_element_by_css_selector(
            '#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_concerne > option:nth-child(1)').click()
        select_postes = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                                             'concerne > option:nth-child(2)')
        if solo_arquetas == 'gc+p':
            select_postes = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                                                 'concerne > option:nth-child(4)')
        elif solo_arquetas == 'p':
            select_postes = browser.find_element_by_css_selector('#\/com\:commande\/gcblo\:cont_com\/gcblo\:cde_'
                                                                 'concerne > option:nth-child(3)')
        select_postes.click()

        # Fechas de cpl y str
        time.sleep(1)
        diferencial_tipo = 'declaration_trvx'
        if dos_type == 'CPL':
            diferencial_tipo = 'contenu_declaration_travaux'

        form_fecha_ini = browser.find_element_by_css_selector(
            '#\/com\:commande\/gcblo\:' + diferencial_tipo + '\/gcblo\:date_debut_trvx')
        form_fecha_fin = browser.find_element_by_css_selector(
            '#\/com\:commande\/gcblo\:' + diferencial_tipo + '\/gcblo\:date_fin_trvx_td_input > input:nth-child(1)')
        time.sleep(1)
        form_fecha_ini.clear()
        form_fecha_ini.send_keys(fecha_ini)
        time.sleep(1)
        form_fecha_fin.clear()
        form_fecha_fin.send_keys(fecha_fin)

        time.sleep(1)
        # Annadimos las calles que pasan por el recorrido
        for i in range(len(calles)):
            calle_form = browser.find_element_by_css_selector(
                '#\/com\:commande\/gcblo\:' + diferencial_tipo + '\/gcblo\:arteres_princ\[' + str(
                    i + 1) + '\]_td_input > input:nth-child(1)')
            time.sleep(1)
            calle_form.clear()
            calle_form.send_keys(calles[i])
            if i < len(calles) - 1:
                time.sleep(1)
                add_button = browser.find_element_by_css_selector(
                    '#\/com\:commande\/gcblo\:' + diferencial_tipo + '\/gcblo\:arteres_princ\[1\]_td_label1 > a:nth-child(1) > img:nth-child(1)')
                add_button.click()

    ref_cli_form = browser.find_element_by_css_selector('#\/com\:commande\/com\:ref_client')
    ref_cli_form.clear()
    ref_cli_form.send_keys(ref_cli)
    time.sleep(1)
    # Click aceptar para sacar el fci
    browser.find_element_by_css_selector('a.sfci_blackb:nth-child(11)').click()
    time.sleep(8)
    try:
        texto = browser.find_element_by_css_selector('#nomerror').text
        texto_split = texto.split(sep='N')
        c = texto_split[1]
        fci = c[2:14]
    except:
        raise Exception(browser.find_element_by_css_selector('span.sfci_error:nth-child(2)').text)
    # fci = 'f123451234512'
    if 'fci_anterior' in d:
        d['fci_compuesto'] = fci + '\n' + d['fci_anterior']
        d['fci'] = fci
    else:
        d['fci'] = fci
    time.sleep(4)


def tsp_operations_1(dosier, ws):
    # operaciones despues de obtener FCI en el tsp
    if 'fci_compuesto' in dosier:
        ws.Cells(dosier['row'], 40).Value = dosier['fci_compuesto']
    else:
        ws.Cells(dosier['row'], 40).Value = dosier['fci']
    if dosier['tipo'] == 'SPL':
        # if 'v' in ws.Cells(dosier['row'], 65).Value:
        ws.Cells(dosier['row'], 65).Value = 'v1'
        ws.Cells(dosier['row'], 66).Value = 'Attente Input TFX'
    ws.Cells(dosier['row'], 47).Value = 'APP'


def tsp_operations_2(dosier, ws):
    # Operaciones despues de hacer el primer depósito
    ws.Cells(dosier['row'], 49).Value = 'v1'
    ws.Cells(dosier['row'], 50).Value = 'DÉPOSÉ'
    ws.Cells(dosier['row'], 51).Value = 'v1'
    ws.Cells(dosier['row'], 52).Value = 'Déposée  en cours'
    today = datetime.today().strftime("%d-%m-%y")
    ws.Cells(dosier['row'], 53).Value = today


def tsp_operations_3(row, ws, v):
    # Operaciones después de hacer el segundo depósito
    ws.Cells(row, 66).Value = 'DÉPOSÉ'
    today = datetime.today().strftime("%d-%m-%y")
    ws.Cells(row, 67).Value = today
    ws.Cells(row, 68).Value = 'v' + v
    ws.Cells(row, 69).Value = 'Déposée  en cours'


def change_c3a(d, fci, client):
    print('Cambiando c3a')
    pythoncom.CoInitialize()
    global destino
    dosier_folder = d[:3] + ' ' + d[3:]
    ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\03-CTRL OK-A depositar' + os.sep
    try:
        for e in os.listdir(ruta):
            if dosier_folder in e:
                dosier_folder = ruta + e + os.sep
        for e in os.listdir(dosier_folder):
            if '_V' in e and 'RF' in e:
                destino = dosier_folder + e + os.sep
        excel_file = destino + 'Fxxxxxxxxxxx_C3A.xls'
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb2 = excel.Workbooks.Open(excel_file)
        # excel.Visible = True
        ws2 = wb2.Worksheets('Commandes Fermes')
        time.sleep(5)
        ws2.Cells(8, 3).Value = fci
        wb2.Close(True)
        new_file = destino + fci + '_C3A.xls'
        os.rename(excel_file, new_file)
    except:
        print('error al modificar la c3a')


def change_dxf(d, fci, client):
    # Llama a una macro de excel para que cambie el nombre de la capa y del bloque de texto en el dxf
    print('Cambiando dxf')
    pythoncom.CoInitialize()
    global destino, dxf_file
    # Buscar ruta del archivo dxf
    dosier_folder = d[:3] + ' ' + d[3:]
    ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\03-CTRL OK-A depositar' + os.sep
    for e in os.listdir(ruta):
        if dosier_folder in e:
            dosier_folder = ruta + e + os.sep
    for e in os.listdir(dosier_folder):
        if '_V' in e and 'RF' in e:
            destino = dosier_folder + e
    for e in os.listdir(destino):
        if '_xxxxx' in e:
            dxf_file = e

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Application.Visible = True
    wb3 = excel.Workbooks.Open(r'C:/Users/josko/PycharmProjects/josko/macro.xlsm')
    ws3 = wb3.Worksheets('DXF')
    ws3.Cells(10, 3).Value = dxf_file
    ws3.Cells(12, 3).Value = destino
    ws3.Cells(14, 3).Value = fci[1:]

    excel.Application.Run('macro.xlsm!OPGC_DXF_UPD')
    wb3.Close(True)
    for e in os.listdir(destino):
        if 'Fxxxxxxxx' in e:
            os.remove(destino + os.sep + e)


def mover_ficheros(d, client, dep=1):
    # Despues de realizar el depósito mueve el dosier al estatuto correspondiente
    print('Moviendo la carpeta')
    dosier_folder = d[:3] + ' ' + d[3:]
    ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\03-CTRL OK-A depositar' + os.sep
    if dep == 2:
        ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\06.9-CTRL OK-TFX a Depositar' + \
               os.sep
    for e in os.listdir(ruta):
        if dosier_folder in e:
            dosier_folder = ruta + e
    origen = dosier_folder
    destino = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\04-Anexos 1 Depositado' + os.sep
    if dep == 2:
        destino = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\07-TFX Depositados' + os.sep
    if os.path.exists(origen):
        ruta = shutil.move(origen, destino)
        print('El directorio ha sido movido a', ruta)
    else:
        print('El directorio origen no existe')


def zip_ficheros(d, fci, client, dep=1):
    # Crea un zip del dxf y el c3a, y las carpetas FOA si hubiera
    global destino, folder_rf
    print('Comprimiendo en zip')
    dosier_folder = d[:3] + ' ' + d[3:]
    ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\03-CTRL OK-A depositar' + os.sep
    if dep == 2:
        ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\06.9-CTRL OK-TFX a Depositar' + \
               os.sep
    # destino = ruta + '04-Anexos 1 Depositado' + os.sep

    for e in os.listdir(ruta):
        if dosier_folder in e:
            dosier_folder = ruta + e + os.sep

    if dep == 2:
        for e in os.listdir(dosier_folder):
            if 'TFX_V' in e:
                destino = dosier_folder + e + os.sep
                folder_rf = e
    else:
        for e in os.listdir(dosier_folder):
            if '_V' in e and 'RF' in e:
                destino = dosier_folder + e + os.sep
                folder_rf = e

    # ZIP DEFLATED para hacer que se comprima, no solo que se guarde en un archivo .zip
    zf = zipfile.ZipFile(destino + folder_rf + '.zip', mode='w', compression=zipfile.ZIP_DEFLATED)
    for e in os.listdir(destino):
        if e != folder_rf + '.zip' and 'xxxxxxxx' not in e:
            zf.write(destino + e, basename(destino + e))
    zf.close()


def depositar_webop(d, fci, browser, client, dep=1):
    # Entra en la página para depositar y sube el ZIP del dosier
    print('Deposito webop')
    global file_path, destino
    dosier_folder = d[:3] + ' ' + d[3:]
    browser.get('https://espaceclient.orange-business.com/group/divop/historique-fci')
    time.sleep(2)
    main_window_handle = browser.current_window_handle
    browser.switch_to_frame(browser.find_element_by_id('ece_iframe'))
    fci_form = browser.find_element_by_xpath('/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/'
                                             'form/table[2]/tbody/tr/td/table[2]/tbody/tr[1]/td/table/tbody/tr'
                                             '[3]/td[2]/input')
    fci_form.send_keys(fci)
    time.sleep(1)
    browser.find_element_by_xpath('/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/'
                                  'tbody/tr/td/table[2]/tbody/tr[1]/td/table/tbody/tr[9]/td[2]/input[1]').clear()
    time.sleep(1)
    browser.find_element_by_xpath('/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/'
                                  'tbody/tr/td/table[2]/tbody/tr[2]/td/a').click()
    time.sleep(2)
    browser.find_element_by_xpath('/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/span/'
                                  'table/tbody/tr[1]/td[1]/a').click()
    time.sleep(2)
    browser.find_element_by_xpath('/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[3]'
                                  '/tbody/tr[1]/td/table/tbody/tr/td[1]/a[3]').click()
    time.sleep(10)
    browser.switch_to_window(browser.window_handles[1])
    browser.refresh()
    time.sleep(10)

    ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\03-CTRL OK-A depositar' + os.sep
    if dep == 2:
        ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\06.9-CTRL OK-TFX a Depositar' + \
               os.sep
    for e in os.listdir(ruta):
        if dosier_folder in e:
            dosier_folder = ruta + e + os.sep
    if dep == 2:
        for e in os.listdir(dosier_folder):
            if 'TFX_V' in e:
                destino = dosier_folder + e + os.sep
    else:
        for e in os.listdir(dosier_folder):
            if '_V' in e and 'RF' in e:
                destino = dosier_folder + e + os.sep

    for e in os.listdir(destino):
        if '.zip' in e:
            file_path = destino + e

            #   Repite este bucle porque a veces no carga bien la página y no sube bien el archivo
    for i in range(7):
        try:
            input_file = browser.find_element_by_id('ufile')
            input_file.send_keys(file_path)
            break
        except:
            browser.refresh()
            time.sleep(8)

    time.sleep(3)
    # boton de subir
    browser.find_element_by_xpath('/html/body/div/div[3]/div/div/div[2]/div[2]/div/div/table/tbody/tr[2]/td[7'
                                  ']/button[1]').click()
    time.sleep(8)
    #   Crear captura de pantalla
    browser.get_screenshot_as_file(destino + 'capt upload ' + fci + '.png')


def descargar_zip_tfx(d, fci, browser, client):
    time.sleep(2)
    browser.get('https://espaceclient.orange-business.com/group/divop/visualiser-vos-fichiers-deposes')
    time.sleep(7)
    WebDriverWait(browser, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'ece_iframe')))
    WebDriverWait(browser, 10).until\
        (EC.presence_of_element_located((By.XPATH, '/html/body/div/div[3]/div[2]/table/tbody/tr[1]/th[7]/input')))
    time.sleep(3)
    fci_form = browser.find_element_by_xpath('/html/body/div/div[3]/div[2]/table/tbody/tr[1]/th[7]/input')
    fci_form.send_keys('%' + fci)
    time.sleep(4)
    WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), '_" + fci + "')]")))
    browser.find_element_by_xpath('/html/body/div/div[3]/div[2]/table/tbody/tr[2]/td[1]/button').click()
    time.sleep(8)


def mover_zip_descargado(fci, client):
    origen = None
    print('Moviendo el zip')
    ruta = os.getcwd() + os.sep
    for i in range(10):
        for e in os.listdir(ruta):
            if fci in e and '.part' not in e:
                origen = ruta + e
                break
        if origen != None:
            break
        else:
            time.sleep(3)
    destino = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\11- GCB1_TFX\\ENVIAR' + os.sep
    if os.path.exists(origen):
        ruta = shutil.move(origen, destino)
        print('El fichero ha sido movido a', ruta)
    else:
        print('El fichero zip no existe')

###################################### COMIENZA EL PROCESO ##################################################

def obtener_fci(client):
    # Carga los datos del excel y se mete en WebOP para obtener el número FCI, si es STR o CPL realiza el primer depósito
    global wb
    logging.basicConfig(filename='webop.log', level=logging.INFO,
                        format='%(asctime)s %(levelname)s: %(message)s', datefmt='%d/%m/%Y %I:%M:%S %p')

    # dosieres de prueba para no tener que cargar el excel continuamente
    # dosieres = {
    #             'Ali195': {'fci': 'F31344040417', 'row': 91, 'es_aval': True, 'IPE_PM': 'FI-92073-0023', 'ref_1era_PM': 'A DEPOSER', 'date_fin': '18/04/2017', 'calles': ['rue del percebe', 'calle street', 'callejon hammer'], 'ref_cli': '', 'solo_arquetas': True, 'ciudad': 'SURESNES', 'date_ini': '28/12/2016', 'nombre': 'Sus1314', 'es_1ca': True, 'cliente': 'SC1', 'tipo': 'STR', 'num_EL': 38}
    # }
    # dosieres = {
    #     'Sal77': {'fci': 'F05920220317', 'formulario': 'SCOP_RFA_TLS_31555_V3', 'nombre': 'Sal77', 'cliente': 'SC00',
    #               'es_aval': True, 'solo_arquetas': 'gc+p', 'otras_ciudades': None, 'date_fin': '31/07/2017',
    #               'ref_cli': 'SC6_TLS_AVALPMRCEM_FI_31555_0111', 'ref_1era_PM': 'A DEPOSER', 'ciudad': 'TOULOUSE',
    #               'tipo': 'CPL', 'IPE_PM': 'FI-31555-0111', 'calles': ['ROUTE DE LAUNAGUET', 'RUE PAGES'], 'row': 55,
    #               'num_EL': '300', 'date_ini': '11/04/2017', 'es_1ca': True}}

    result = {}
    try:
        dosieres = cargar_datos_excel(client)
        pass
    except Exception as ex:
        logging.error('No han podido cargarse los datos del excel porque: %s', ex.args[0])
        result[ex.args[1]] = ex.args[1] + ' ' + ex.args[0]
    pythoncom.CoInitialize()

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    # excel.DisplayAlerts = True
    excel.EnableEvents = False
    if os.name == 'nt':
        if client == 'SC1':
            wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC1/PRODUCCIÓN/Tab Suivi Prod/SC1 TSP 2017.xlsm')
        elif client == 'SC00':
            wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC00/PRODUCCIÓN/Tab Suivi Prod/SC00 TSP 2017.xlsm')
        elif client == 'SC6':
            wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC6/PRODUCCIÓN/Tab Suivi Prod/SC6 TSP 2017.xlsm')
        elif client == 'SC4':
            wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC4/PRODUCCIÓN/Tab Suivi Prod/SC4 TSP 2017.xlsm')
            # wb = excel.Workbooks.Open(r'C:/Users/josko/PycharmProjects/josko/SuiviJRU.xlsm')
    else:
        wb = excel.Workbooks.Open(r'\home\ubuntu\3id2plus\SuiviJRU.xlsx')
    excel.Visible = False
    ws = wb.Worksheets('Tab Suivi Prod')
    time.sleep(5)

    for d in dosieres:
        try:
            # fci = 'F28988160217'
            # dosieres[d]['fci'] = fci
            browser = set_up_browser()
            if dosieres[d]['otras_ciudades'] is not None:
                raise Exception('Tiene más de una ciudad')
            login(browser, client)
            boutique_operations(browser, dosieres[d])
            time.sleep(4)
            tsp_operations_1(dosieres[d], ws)
            if dosieres[d]['tipo'] == 'CPL' or dosieres[d]['tipo'] == 'STR':
                change_c3a(d, dosieres[d]['fci'], client)
                change_dxf(d, dosieres[d]['fci'], client)
                zip_ficheros(d, dosieres[d]['fci'], client)
                depositar_webop(d, dosieres[d]['fci'], browser, client)
                mover_ficheros(d, client)
                tsp_operations_2(dosieres[d], ws)

        except Exception as ex:
            try:
                logging.error('%s No ha podido completarse por: %s', dosieres[d]['nombre'], ex.args[0])
                result[dosieres[d]['nombre']] = dosieres[d]['nombre'] + ' No ha podido completarse por: ' + ex.args[0]
            except Exception:
                print('Error de argumentos en la excepción')

        else:
            logging.info('%s --> Se ha procesado correctamente: ', dosieres[d]['nombre'])
            result[dosieres[d]['nombre']] = dosieres[d]['nombre'] + '  --> Se ha procesado correctamente: '

        finally:
            if os.name != 'nt':
                browser.service.process.send_signal(signal.SIGTERM)
            browser.quit()
            time.sleep(5)
    print('Salvando excel')
    wb.Close(True)
    # wb.SaveAs(r'C:\Users\josko\PycharmProjects\josko\SuiviJRU3.xlsx')
    excel.Application.Quit()
    return result


def depositar2(client):

    # Realiza el segundo depósito de los dosieres que estan en la carpeta 6.9
    print('segundo depósito')
    global destino
    # dosieres = {
    #     'Sal77': {'fci': 'F05920220317', 'formulario': 'SCOP_RFA_TLS_31555_V3', 'nombre': 'Sal77', 'cliente': 'SC00',
    #               'es_aval': True, 'solo_arquetas': 'gc+p', 'otras_ciudades': None, 'date_fin': '31/07/2017',
    #               'ref_cli': 'SC6_TLS_AVALPMRCEM_FI_31555_0111', 'ref_1era_PM': 'A DEPOSER', 'ciudad': 'TOULOUSE',
    #               'tipo': 'CPL', 'IPE_PM': 'FI-31555-0111', 'calles': ['ROUTE DE LAUNAGUET', 'RUE PAGES'], 'row': 55,
    #               'num_EL': '300', 'date_ini': '11/04/2017', 'es_1ca': True}}
    # for d in dosieres:
    #     browser = set_up_browser()
    #     login(browser, client)
    #     descargar_zip_tfx(d, dosieres[d]['fci'], browser, client)
    #     mover_zip_descargado(dosieres[d]['fci'], client)

    result = {}
    version = 1
    ruta = 'Z:\\03-PRODUCCION\\0.CAFT\\' + client + '\\PRODUCCIÓN\\PROD_Interna\\06.9-CTRL OK-TFX a Depositar' + os.sep
    dosieres = {}
    # dosieres = {'Sal77': {'fci': 'F05920220317', 'formulario': 'SCOP_RFA_TLS_31555_V3', 'nombre': 'Sal77', 'cliente': 'SC00', 'es_aval': True, 'solo_arquetas': 'gc+p', 'otras_ciudades': None, 'date_fin': '31/07/2017', 'ref_cli': 'SC6_TLS_AVALPMRCEM_FI_31555_0111', 'ref_1era_PM': 'A DEPOSER', 'ciudad': 'TOULOUSE', 'tipo': 'CPL', 'IPE_PM': 'FI-31555-0111', 'calles': ['ROUTE DE LAUNAGUET', 'RUE PAGES'], 'row': 55, 'num_EL': '300', 'date_ini': '11/04/2017', 'es_1ca': True}}
    for e in os.listdir(ruta):
        nom_dossier = e.split(sep=' ')
        dossier = nom_dossier[0] + nom_dossier[1]
        dosieres[dossier] = {}
        dossier_folder = ruta + e + os.sep
        for f in os.listdir(dossier_folder):
            if 'TFX_V' in f:
                destino = dossier_folder + f + os.sep
                version = f[5]
        for f in os.listdir(destino):
            if 'C3' in f:
                fci = f[:12]
                dosieres[dossier]['fci'] = fci

    pythoncom.CoInitialize()
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.EnableEvents = False
    if client == 'SC1':
        wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC1/PRODUCCIÓN/Tab Suivi Prod/SC1 TSP 2017.xlsm')
    elif client == 'SC00':
        wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC00/PRODUCCIÓN/Tab Suivi Prod/SC00 TSP 2017.xlsm')
    elif client == 'SC6':
        wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC6/PRODUCCIÓN/Tab Suivi Prod/SC6 TSP 2017.xlsm')
    elif client == 'SC4':
        wb = excel.Workbooks.Open(r'Z:/03-PRODUCCION/0.CAFT/SC4/PRODUCCIÓN/Tab Suivi Prod/SC4 TSP 2017.xlsm')
    ws = wb.Worksheets('Tab Suivi Prod')
    for d in dosieres:
        print('Segundo depósito de ' + d)
        browser = set_up_browser()
        login(browser, client)

        try:
            zip_ficheros(d, dosieres[d]['fci'], client, 2)
            depositar_webop(d, dosieres[d]['fci'], browser, client, 2)
        except:
            result[d] = 'El dosier ' + d + ' no se ha depositado correctamente'
        else:
            try:
                descargar_zip_tfx(d, dosieres[d]['fci'], browser, client)
                mover_zip_descargado(dosieres[d]['fci'], client)
                pass
            except:
                result[d] = 'En el dosier ' + d + ' no se ha descargado el .zip o no se ha movido a la carpeta'
            else:
                try:
                    mover_ficheros(d, client, 2)
                except:
                    result[d] = 'El dosier ' + d + ' ha fallado al mover las carpetas'
                else:
                    for i in range(9, 1500):
                        if ws.Cells(i, 1).GetValue() == d:
                            dosieres[d]['row'] = i
                            break
                    try:
                        tsp_operations_3(dosieres[d]['row'], ws, version)
                        pass
                    except:
                        result[d] = 'El dosier ' + d + ' se ha depositado bien pero no se ha encontrado en TSP 2017'
                    else:
                        result[d] = 'El dosier ' + d + ' se ha procesado correctamente'
        finally:
            browser.quit()
            time.sleep(5)
    print('Salvando excel')
    wb.Close(True)
    excel.Application.Quit()
    return result
